import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.dateparse import parse_date

from apps.proveedores.models import IngresoMateriaPrima, Proveedor
from apps.produccion.models import Paca
from apps.ventas.models import Venta
from apps.despachos.models import Despacho
from apps.clientes.models import Cliente
from apps.transportes.models import Transporte
from apps.accounts.models import PerfilUsuario, Rol  # <--- Corregido a 'accounts'

@staff_member_required
def panel_reportes(request):
    return render(request, 'reportes/panel.html')


def _obtener_fechas_filtro(request):
    """Función auxiliar para extraer y validar fechas 'desde' y 'hasta'."""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    inicio = parse_date(fecha_inicio) if fecha_inicio else None
    fin = parse_date(fecha_fin) if fecha_fin else None
    return inicio, fin


# ==================== 1. CLIENTES ====================
@staff_member_required
def reporte_clientes(request):
    clientes = Cliente.objects.all().order_by('nombre')
    total_clientes = clientes.count()
    return render(request, 'reportes/reporte_clientes.html', {'clientes': clientes, 'total_clientes': total_clientes})

@staff_member_required
def exportar_clientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    ws.append(['ID', 'Nombre', 'RUC / Cédula', 'Teléfono', 'Correo', 'Dirección', 'Estado', 'Fecha Registro'])
    for c in Cliente.objects.all().order_by('nombre'):
        ws.append([c.id, c.nombre, c.ruc, c.telefono, c.correo, c.direccion, "Activo" if c.activo else "Inactivo", str(c.fecha_registro)])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_clientes.xlsx'
    wb.save(response)
    return response


# ==================== 2. PROVEEDORES ====================
@staff_member_required
def reporte_proveedores(request):
    proveedores = Proveedor.objects.all().order_by('nombre')
    total_proveedores = proveedores.count()
    return render(request, 'reportes/reporte_proveedores.html', {'proveedores': proveedores, 'total_proveedores': total_proveedores})

@staff_member_required
def exportar_proveedores_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    ws.append(['ID', 'Nombre', 'RUC', 'Teléfono', 'Correo', 'Dirección', 'Estado'])
    for p in Proveedor.objects.all().order_by('nombre'):
        ws.append([p.id, p.nombre, p.ruc, p.telefono, p.correo, p.direccion, "Activo" if p.activo else "Inactivo"])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_proveedores.xlsx'
    wb.save(response)
    return response


# ==================== 3. TRANSPORTISTAS / CHOFERES ====================
@staff_member_required
def reporte_transportes(request):
    transportes = Transporte.objects.all().order_by('placa')
    total_transportes = transportes.count()
    return render(request, 'reportes/reporte_transportes.html', {'transportes': transportes, 'total_transportes': total_transportes})

@staff_member_required
def exportar_transportes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transportistas"
    ws.append(['ID', 'Placa', 'Identificación / Cédula', 'Propietario', 'Conductor', 'Teléfono', 'Capacidad (Kg)', 'Estado'])
    for t in Transporte.objects.all().order_by('placa'):
        ws.append([t.id, t.placa, t.identificacion, t.propietario, t.conductor, t.telefono, float(t.capacidad_kg), "Activo" if t.activo else "Inactivo"])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_transportes.xlsx'
    wb.save(response)
    return response


# ==================== 4. MATERIA PRIMA ====================
@staff_member_required
def reporte_materia_prima(request):
    inicio, fin = _obtener_fechas_filtro(request)
    ingresos = IngresoMateriaPrima.objects.all()
    if inicio:
        ingresos = ingresos.filter(fecha__gte=inicio)
    if fin:
        ingresos = ingresos.filter(fecha__lte=fin)
    ingresos = ingresos.order_by('-id')
    total_kg = sum(item.peso_neto for item in ingresos)
    
    context = {'ingresos': ingresos, 'total_kg': total_kg, 'fecha_inicio': request.GET.get('fecha_inicio', ''), 'fecha_fin': request.GET.get('fecha_fin', '')}
    return render(request, 'reportes/reporte_materia_prima.html', context)

@staff_member_required
def exportar_materia_prima_excel(request):
    inicio, fin = _obtener_fechas_filtro(request)
    ingresos = IngresoMateriaPrima.objects.all()
    if inicio: ingresos = ingresos.filter(fecha__gte=inicio)
    if fin: ingresos = ingresos.filter(fecha__lte=fin)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingresos Materia Prima"
    ws.append(['ID', 'Fecha', 'Proveedor', 'RUC', 'Tipo de Material', 'Cantidad / Unidades', 'Peso Neto (Kg)', 'Observación'])
    for item in ingresos.order_by('-id'):
        ws.append([item.id, str(item.fecha), item.proveedor.nombre, item.proveedor.ruc, item.get_tipo_material_display(), float(item.cantidad), float(item.peso_neto), item.observacion])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_materia_prima.xlsx'
    wb.save(response)
    return response


# ==================== 5. PRODUCCIÓN ====================
# ==================== 5. PRODUCCIÓN ====================
@staff_member_required
def reporte_produccion(request):
    inicio, fin = _obtener_fechas_filtro(request)
    pacas = Paca.objects.select_related('detalle__produccion__operador', 'detalle__producto', 'detalle__color').all()
    
    if inicio:
        pacas = pacas.filter(detalle__produccion__fecha__gte=inicio)
    if fin:
        pacas = pacas.filter(detalle__produccion__fecha__lte=fin)
        
    # Cambiado a orden ascendente (ID 1 primero)
    pacas = pacas.order_by('id')
    context = {
        'pacas': pacas,
        'fecha_inicio': request.GET.get('fecha_inicio', ''),
        'fecha_fin': request.GET.get('fecha_fin', '')
    }
    return render(request, 'reportes/reporte_produccion.html', context)

@staff_member_required
def exportar_produccion_excel(request):
    inicio, fin = _obtener_fechas_filtro(request)
    pacas = Paca.objects.select_related('detalle__produccion__operador', 'detalle__producto', 'detalle__color').all()
    
    if inicio:
        pacas = pacas.filter(detalle__produccion__fecha__gte=inicio)
    if fin:
        pacas = pacas.filter(detalle__produccion__fecha__lte=fin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Producción de Pacas"
    ws.append(['ID Paca', 'Código Único', 'Fecha', 'Operador', 'Producto', 'Color', 'Lote', 'N° Paca'])
    
    # Cambiado aquí también para que el Excel baje correlativo desde el ID 1
    for p in pacas.order_by('id'):
        prod = p.detalle.produccion
        ws.append([
            p.id,
            p.codigo_unico,
            str(prod.fecha),
            prod.operador.nombre_completo,
            str(p.detalle.producto),
            str(p.detalle.color),
            p.lote,
            p.numero_paca
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_produccion.xlsx'
    wb.save(response)
    return response


# ==================== 6. VENTAS ====================
@staff_member_required
def reporte_ventas(request):
    inicio, fin = _obtener_fechas_filtro(request)
    ventas = Venta.objects.all()
    if inicio:
        ventas = ventas.filter(fecha__gte=inicio)
    if fin:
        ventas = ventas.filter(fecha__lte=fin)
    ventas = ventas.order_by('-fecha')
    total_vendido = sum(v.total for v in ventas)
    
    context = {'ventas': ventas, 'total_vendido': total_vendido, 'fecha_inicio': request.GET.get('fecha_inicio', ''), 'fecha_fin': request.GET.get('fecha_fin', '')}
    return render(request, 'reportes/reporte_ventas.html', context)

@staff_member_required
def exportar_ventas_excel(request):
    inicio, fin = _obtener_fechas_filtro(request)
    ventas = Venta.objects.all()
    if inicio: ventas = ventas.filter(fecha__gte=inicio)
    if fin: ventas = ventas.filter(fecha__lte=fin)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(['N° Factura', 'Fecha', 'Cliente', 'RUC Cliente', 'Total ($)', 'Observación'])
    for v in ventas.order_by('-fecha'):
        ws.append([v.numero_factura, str(v.fecha), v.cliente.nombre, v.cliente.ruc, float(v.total), v.observacion or ''])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_ventas.xlsx'
    wb.save(response)
    return response


# ==================== 7. DESPACHOS ====================
@staff_member_required
def reporte_despachos(request):
    inicio, fin = _obtener_fechas_filtro(request)
    despachos = Despacho.objects.all()
    if inicio:
        despachos = despachos.filter(fecha__gte=inicio)
    if fin:
        despachos = despachos.filter(fecha__lte=fin)
    despachos = despachos.order_by('-fecha')
    
    context = {'despachos': despachos, 'fecha_inicio': request.GET.get('fecha_inicio', ''), 'fecha_fin': request.GET.get('fecha_fin', '')}
    return render(request, 'reportes/reporte_despachos.html', context)

@staff_member_required
def exportar_despachos_excel(request):
    inicio, fin = _obtener_fechas_filtro(request)
    despachos = Despacho.objects.all()
    if inicio: despachos = despachos.filter(fecha__gte=inicio)
    if fin: despachos = despachos.filter(fecha__lte=fin)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Despachos"
    ws.append(['N° Guía', 'Fecha', 'Cliente', 'Transportista (Placa)', 'Conductor', 'Estado', 'Dirección'])
    for d in despachos.order_by('-fecha'):
        ws.append([d.numero_guia, str(d.fecha), d.cliente.nombre, d.transporte.placa, d.transporte.conductor, d.estado, d.direccion_entrega])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_despachos.xlsx'
    wb.save(response)
    return response

# ==================== 6. USUARIOS ====================
@staff_member_required
def reporte_usuarios(request):
    # Obtenemos los perfiles con su usuario y rol optimizados
    perfiles = PerfilUsuario.objects.select_related('usuario', 'rol').all().order_by('id')
    
    context = {
        'perfiles': perfiles,
    }
    return render(request, 'reportes/reporte_usuarios.html', context)

@staff_member_required
def exportar_usuarios_excel(request):
    perfiles = PerfilUsuario.objects.select_related('usuario', 'rol').all().order_by('id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Usuarios"
    ws.append(['ID', 'Usuario (Username)', 'Nombre Completo', 'Correo', 'Código Operador', 'Rol', 'Cargo', 'Teléfono', 'Activo'])
    
    for p in perfiles:
        ws.append([
            p.id,
            p.usuario.username,
            p.usuario.get_full_name() or "N/A",
            p.usuario.email or "N/A",
            p.codigo_operador,
            str(p.rol) if p.rol else "Sin Rol",
            p.cargo or "N/A",
            p.telefono or "N/A",
            "Sí" if p.activo else "No"
        ])
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_usuarios.xlsx'
    wb.save(response)
    return response